"""
KROK 2 - Otisk stranky.

Projde /cz/dokumentace-ke-stazeni, z ni vytahne vsechny kategorie a produkty,
otevre kazdou produktovou stranku a zaznamena VSECHNY dokumenty:
kde jsou (kategorie / produkt / sekce / poradi), jak se jmenuje popis,
jak se jmenuje soubor, jaky ma typ, velikost a poznamku.

Vysledek: exports/snapshot_RRRRMMDD_HHMMSS.xlsx (a/nebo .csv)
Prubeh + pripadne chyby: logs/snapshot_RRRRMMDD_HHMMSS.log

Spusteni:
    python snapshot.py                    # normalne (potrebuje session.json)
    python snapshot.py --headed           # s viditelnym prohlizecem
    python snapshot.py --limit 3          # jen prvni 3 produkty (test)
    python snapshot.py --format both      # xlsx i csv
    python snapshot.py --out muj_nazev    # vlastni zaklad nazvu vystupu
    python snapshot.py --types pdf,dwg    # jen dane typy souboru
    python snapshot.py --category "vzduch/voda" --brand ivt  # filtr produktu
    python snapshot.py --section "technicke listy"           # filtr sekci
    python snapshot.py --dry-run          # jen souhrn, nic se nezapise
    python snapshot.py --nejvyhledavanejsi  # zaznamenat i "Nejvyhledavanejsi dokumenty"
    python snapshot.py --no-dwg           # nezaznamenavat DWG soubory
"""

import argparse
import csv
import datetime as dt
import difflib
import logging
import sys
import time
import unicodedata
from pathlib import Path
from urllib.parse import urljoin

from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from extractor import (
    JS_EXTRACT_TABLES, header_index, cell, find_size,
    is_locked, is_download, file_path_from_href, check_session,
    GREEN, RESET, load_config, cfg_bool,
)

BASE = "https://www.projektuj-tepelna-cerpadla.cz"
DOCS_URL = f"{BASE}/cz/dokumentace-ke-stazeni"
HERE = Path(__file__).parent
SESSION_FILE = HERE / "session.json"
EXPORT_DIR = HERE / "exports"
LOG_DIR = HERE / "logs"

COLUMNS = [
    "kategorie",
    "znacka",
    "typ_produktu",
    "sekce",
    "poradi_v_sekci",
    "popis_dokumentu",
    "typ_souboru",
    "nazev_souboru",
    "velikost",
    "poznamka",
    "url_stazeni",
]
# "produkt", "dostupnost", "produkt_url", "cesta_souboru" a "klic" se pocitaji
# dal (viz scrape_product), ale do xlsx/csv se uz nezapisuji - nejsou v COLUMNS.
# Identitu radku (kategorie+znacka+typ_produktu+sekce+poradi_v_sekci) si tedy
# musi poskladat sam kdokoliv, kdo bude otisky dal zpracovavat/porovnavat.


class _ConsoleFilter(logging.Filter):
    """Vynecha z konzole zpravy oznacene extra={'file_only': True} - pouzivaji
    se pro detailni radek za kazdy produkt, ktery by jinak rozbijel progress bar."""

    def filter(self, record):
        return not getattr(record, "file_only", False)


def setup_logging(run_id):
    """Zaloguje beh do logs/snapshot_<run_id>.log a soucasne na konzoli.
    Vyjimky se zapisuji s celym tracebackem, aby slo dohledat, kde presne
    beh spadl."""
    LOG_DIR.mkdir(exist_ok=True)
    log_path = LOG_DIR / f"snapshot_{run_id}.log"

    logger = logging.getLogger("snapshot")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%Y-%m-%d %H:%M:%S")

    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(logging.Formatter("%(message)s"))  # konzole beze zmen puvodniho vzhledu
    ch.addFilter(_ConsoleFilter())
    logger.addHandler(ch)

    return logger, log_path


class ProgressBar:
    """Jednoradkovy progress bar v konzoli (prepisuje se pres \\r), nezavisly
    na logovani do souboru - detailni radky za produkt jdou porad do log
    souboru (viz _ConsoleFilter), tohle je jen vizualni ukazatel postupu."""

    WIDTH = 28

    def __init__(self):
        self._last_len = 0

    def update(self, i, total, extra=""):
        if total <= 0 or not sys.stdout.isatty():
            return
        filled = int(self.WIDTH * i / total)
        bar = f"{GREEN}{'#' * filled}{RESET}" + "-" * (self.WIDTH - filled)
        pct = int(100 * i / total)
        text = f"\r  [{bar}] {pct:3d}%  {i}/{total}  {extra}"
        pad = max(0, self._last_len - len(text))
        sys.stdout.write(text + " " * pad)
        sys.stdout.flush()
        self._last_len = len(text)

    def interrupt(self):
        """Zavola se pred logovanim neceho jineho na konzoli (warning/error),
        aby se progress bar neprepsal napulku novou zpravou."""
        if self._last_len:
            sys.stdout.write("\n")
            sys.stdout.flush()
            self._last_len = 0

    def finish(self):
        if self._last_len:
            sys.stdout.write("\n")
            sys.stdout.flush()
            self._last_len = 0


# ---------------------------------------------------------------------------
# 1) Rozcestnik - kategorie + produkty
# ---------------------------------------------------------------------------
# JS: projde DOM a k  odkazu "Vybrat" najde znacku+typ ze sousednich bunek
# a kategorii z nejblizsiho predchoziho nadpisu/odstavce mimo tabulku.
JS_INDEX = r"""
() => {
  const BASE = 'https://www.projektuj-tepelna-cerpadla.cz';
  const SKIP = ['prihlaseni','registrace','kontakt','kontaktni-formular',
                'vyhledavani','prehled','jak-mam','chyby','vzorove',
                'skoleni','o-nas','download='];

  function clean(s){ return (s||'').replace(/\s+/g,' ').trim(); }

  // Kategorie: nejblizsi predek s tridou "toggle-block", jeho primy potomek
  // ".toggle-block-title" (vizualne vypada jako <h2>, ale je to <div>).
  function findCategory(row){
    const block = row.closest('.toggle-block');
    if(!block) return '';
    const titleEl = block.querySelector(':scope > .toggle-block-title');
    if(!titleEl) return '';
    const textEl = titleEl.querySelector('.toggle-block-title-text') || titleEl;
    return clean(textEl.innerText);
  }

  const results = [];
  const seen = new Set();

  document.querySelectorAll('a').forEach(a => {
    const href = a.getAttribute('href') || '';
    const text = clean(a.innerText);
    if(text !== 'Vybrat') return;
    if(!href.startsWith('/cz/') && !href.startsWith(BASE+'/cz/')) return;
    const abs = href.startsWith('http') ? href : BASE+href;
    if(seen.has(abs)) return;
    if(SKIP.some(s => abs.includes(s))) return;
    seen.add(abs);

    const row = a.closest('.toggle-block-trow');
    const cat = row ? findCategory(row) : '';

    // znacka a typ: primo v radku jsou divy s temito tridami
    let znacka = '', typ = '';
    if(row){
      const brandEl = row.querySelector('.brand');
      const typeEl = row.querySelector('.type');
      znacka = brandEl ? clean(brandEl.innerText) : '';
      typ = typeEl ? clean(typeEl.innerText) : '';
    }

    results.push({ url: abs, kategorie: cat, znacka, typ });
  });

  return results;
}
"""


def scrape_index(page):
    page.goto(DOCS_URL, wait_until="domcontentloaded")
    page.wait_for_timeout(1200)

    items = page.evaluate(JS_INDEX)

    products, seen = [], set()
    for item in items:
        url = item.get("url", "")
        if not url or url in seen:
            continue
        seen.add(url)
        products.append({
            "kategorie": item.get("kategorie", ""),
            "znacka": item.get("znacka", ""),
            "typ_produktu": item.get("typ", ""),
            "produkt_url": url,
        })

    return products


# ---------------------------------------------------------------------------
# 2) Produktova stranka - dokumenty
# ---------------------------------------------------------------------------
JS_PRODUCT = r"""
() => {
  const clean = s => (s || '').replace(/\s+/g,' ').trim();
  const BASE = 'https://www.projektuj-tepelna-cerpadla.cz';

  // Sekce: nejblizsi predek s tridou "toggle-block", jeho primy potomek
  // ".toggle-block-title" (vizualne vypada jako <h2>, ale je to <div>,
  // takze hledani skutecnych H2-H5 tagu ho minne).
  function getSekce(row) {
    const block = row.closest('.toggle-block');
    if (!block) return '(bez sekce)';
    const titleEl = block.querySelector(':scope > .toggle-block-title');
    if (!titleEl) return '(bez sekce)';
    const textEl = titleEl.querySelector('.toggle-block-title-text') || titleEl;
    const t = clean(textEl.innerText);
    return t || '(bez sekce)';
  }

  const h1 = clean((document.querySelector('h1')||{}).innerText||'');
  const title = clean(document.title);

  const rows = Array.from(document.querySelectorAll('.toggle-block-trow'));
  const results = [];
  const sekcCache = new Map();

  rows.forEach(row => {
    const divs = Array.from(row.children);
    if (divs.length < 2) return;

    const popis = clean(divs[0] ? divs[0].innerText : '');
    if (!popis || popis === 'Popis dokumentu') return;

    const typ = clean(divs[1] ? divs[1].innerText : '');
    const noteDiv = row.querySelector('.note');
    const poznamka = noteDiv ? clean(noteDiv.innerText) : '';
    const radekText = clean(row.innerText);

    // Sekce - cachujeme per parent element
    const parentEl = row.parentElement;
    if (!sekcCache.has(parentEl)) sekcCache.set(parentEl, getSekce(row));
    const sekce = sekcCache.get(parentEl);

    const links = Array.from(row.querySelectorAll('a')).filter(a => {
      const h = a.getAttribute('href') || '';
      return h.includes('download=') || h === 'javascript:;';
    }).filter(a => clean(a.innerText) !== '');  // preskoc prazdne ikonkove odkazy

    if (links.length === 0) {
      // Radek bez odkazu - jen zaznamenaj
      results.push({ sekce, popis, typ, nazev:'', href:'', abs:'', poznamka, radekText, locked: true });
    } else {
      links.forEach(a => {
        const href = a.getAttribute('href') || '';
        results.push({ sekce, popis, typ,
          nazev: clean(a.innerText), href, abs: a.href || '',
          poznamka, radekText, locked: href === 'javascript:;' });
      });
    }
  });

  return { h1, title, rows: results };
}
"""


# Sekce, ktere se do vysledku vubec nezaznamenavaji vzdy (nejsou to skutecne
# kategorie dokumentu, ale marketingovy/informacni blok bez dokumentu).
EXCLUDED_SECTIONS = {
    "Aktuální informace o dostupnosti výrobku",
}

# "Nejvyhledávanější dokumenty" je specialni pripad - jsou to skutecne
# dokumenty, jen duplicitni s tim, co uz je v jine sekci nize na strance
# (technicke listy, navody...). Ve vychozim stavu se proto nezaznamenava,
# aby otisk neobsahoval duplicity - jde zapnout pres --nejvyhledavanejsi.
NEJVYHLEDAVANEJSI_SECTION = "Nejvyhledávanější dokumenty"


def scrape_product(page, product, include_nejvyhledavanejsi=False):
    page.goto(product["produkt_url"], wait_until="domcontentloaded")
    page.wait_for_timeout(800)

    data = page.evaluate(JS_PRODUCT)
    nazev_produktu = data["h1"] or data["title"]
    raw_rows = data["rows"]

    out = []
    grouped = []

    for r in raw_rows:
        sekce = r["sekce"]
        popis = r["popis"]
        if not popis or sekce in EXCLUDED_SECTIONS:
            continue
        if sekce == NEJVYHLEDAVANEJSI_SECTION and not include_nejvyhledavanejsi:
            continue

        # Novy radek = jina sekce NEBO jiny popis
        if not grouped or grouped[-1]["sekce"] != sekce or grouped[-1]["popis"] != popis:
            grouped.append({
                "sekce": sekce, "popis": popis,
                "soubory": [],
                "radekText": r["radekText"],
                "typ": r["typ"],
                "poznamka": r["poznamka"],
            })
        if r["nazev"] or r["href"]:
            grouped[-1]["soubory"].append({"nazev": r["nazev"], "href": r["href"], "abs": r["abs"], "locked": r["locked"]})

    # poradi_v_sekci je cele cislo pocitane per soubor (ne per radek) -
    # radek se dvema soubory (napr. schema PDF+DWG) da dve po sobe jdouci cisla.
    sekce_poradi = {}

    for grp in grouped:
        sekce = grp["sekce"]
        soubory = grp["soubory"]
        if not soubory:
            soubory = [{"nazev": "", "href": "", "abs": "", "locked": True}]

        for s in soubory:
            href = s["href"]
            locked = s["locked"]
            cesta = file_path_from_href(s["abs"] or href)

            typ = grp["typ"]
            if len(soubory) > 1:
                ext = cesta.rsplit(".", 1)[-1].upper() if "." in cesta else ""
                typ = ext or typ

            sekce_poradi[sekce] = sekce_poradi.get(sekce, 0) + 1
            poradi = sekce_poradi[sekce]
            klic = f"{product['produkt_url']}|{sekce}|{poradi}"

            out.append({
                "klic": klic,
                "kategorie": product["kategorie"],
                "znacka": product["znacka"],
                "typ_produktu": product["typ_produktu"],
                "produkt": nazev_produktu,
                "produkt_url": product["produkt_url"],
                "sekce": sekce,
                "poradi_v_sekci": poradi,
                "popis_dokumentu": grp["popis"],
                "typ_souboru": typ,
                "nazev_souboru": s["nazev"],
                "cesta_souboru": cesta,
                "velikost": find_size(grp["radekText"]),
                "poznamka": grp["poznamka"],
                "url_stazeni": "" if locked else (s["abs"] or ""),
                "dostupnost": "zamceno / nedostupne" if locked else "ke stazeni",
            })

    return out




# ---------------------------------------------------------------------------
# 3) Zapis do Excelu
# ---------------------------------------------------------------------------
def write_xlsx(records, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Otisk"

    head_fill = PatternFill("solid", fgColor="1F4E79")
    head_font = Font(color="FFFFFF", bold=True)

    ws.append(COLUMNS)
    for c in range(1, len(COLUMNS) + 1):
        cellx = ws.cell(row=1, column=c)
        cellx.fill = head_fill
        cellx.font = head_font
        cellx.alignment = Alignment(vertical="center")

    for rec in records:
        ws.append([rec.get(c, "") for c in COLUMNS])

    widths = {
        "klic": 55, "kategorie": 45, "znacka": 14, "typ_produktu": 18,
        "produkt": 32, "produkt_url": 52, "sekce": 34, "poradi_v_sekci": 13,
        "popis_dokumentu": 52, "typ_souboru": 11, "nazev_souboru": 42,
        "cesta_souboru": 48, "velikost": 10, "poznamka": 24, "dostupnost": 20,
        "url_stazeni": 60,
    }
    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 18)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"

    # info list
    info = wb.create_sheet("Info")
    info.append(["Vytvoreno", dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    info.append(["Zdroj", DOCS_URL])
    info.append(["Poradku celkem", len(records)])
    info.append(["Produktu", len({r["produkt_url"] for r in records})])
    info.append(["Zamcenych", sum(1 for r in records if r["dostupnost"] != "ke stazeni")])
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 60

    wb.save(path)


def write_csv(records, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(COLUMNS)
        for rec in records:
            writer.writerow([rec.get(c, "") for c in COLUMNS])


# ---------------------------------------------------------------------------
def main():
    # Banner se tady nevypisuje - ukazuje se jen jednou pri "nahozeni" (bare
    # "python protc_otisk.py" / START-mac.command), ne pri kazdem jednotlivem
    # prikazu, viz protc_otisk.py.

    # Vychozi hodnoty se berou z config.txt (kdyz existuje a klic je vyplneny),
    # jinak se pouziji vestavene defaulty nize. Cokoliv zadane primo v prikazove
    # radce config.txt pro dany beh prebije - je to jen vychozi hodnota.
    cfg = load_config()

    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=None,
                     help="zakladni nazev vystupu (bez pripony) - podle --format "
                          "se pripoji .xlsx a/nebo .csv")
    ap.add_argument("--format", choices=["xlsx", "csv", "both"],
                     default=cfg.get("format") or "xlsx",
                     help="format vystupu (vychozi podle config.txt)")
    ap.add_argument("--headed", action="store_true",
                     default=cfg_bool(cfg, "headed"),
                     help="zobrazit prohlizec")
    ap.add_argument("--limit", type=int, default=0, help="max. poctu produktu (test)")
    ap.add_argument("--no-session", action="store_true",
                     default=cfg_bool(cfg, "no_session"),
                     help="bez prihlaseni (jen verejna cast)")
    ap.add_argument("--types", default=cfg.get("types") or "all",
                     help="ktere typy souboru ulozit, oddelene carkou (napr. 'pdf' nebo "
                          "'pdf,dwg'). Vychozi podle config.txt")
    ap.add_argument("--category", default=cfg.get("category") or None,
                     help="filtr podle kategorie na rozcestniku - substring, "
                          "bez ohledu na velikost pismen i diakritiku, carkou "
                          "oddelene varianty (napr. 'vzduch/voda')")
    ap.add_argument("--brand", default=cfg.get("brand") or None,
                     help="filtr podle znacky - substring, bez ohledu na velikost "
                          "pismen i diakritiku, carkou oddelene varianty")
    ap.add_argument("--product-type", default=cfg.get("product_type") or None, dest="product_type",
                     help="filtr podle typu produktu - substring, bez ohledu na "
                          "velikost pismen i diakritiku, carkou oddelene varianty")
    ap.add_argument("--section", default=cfg.get("section") or None,
                     help="filtr podle sekce dokumentu - substring, bez ohledu na "
                          "velikost pismen i diakritiku, carkou oddelene varianty "
                          "(napr. 'technicke listy,navody'). Na rozdil od "
                          "--category/--brand/--product-type se aplikuje az po "
                          "nacteni produktovych stranek, protoze sekce se pozna az tam")
    ap.add_argument("--dry-run", action="store_true", dest="dry_run",
                     default=cfg_bool(cfg, "dry_run"),
                     help="provede cely scrape, ale nic nezapise - jen ukaze souhrn")
    ap.add_argument("--nejvyhledavanejsi", action="store_true",
                     default=cfg_bool(cfg, "nejvyhledavanejsi"),
                     help="zaznamenat i sekci 'Nejvyhledavanejsi dokumenty' "
                          "(vychozi vypnuto - jsou to duplicity dokumentu, "
                          "ktere uz jsou v jinych sekcich nize na strance)")
    ap.add_argument("--no-dwg", action="store_true", dest="no_dwg",
                     default=cfg_bool(cfg, "no_dwg"),
                     help="nezaznamenavat DWG soubory (vychozi zaznamenavat) - "
                          "zkratka za '--types' bez 'dwg', nezavisla na tom, "
                          "co je zadane v --types")

    # parse_known_args misto parse_args, abychom u preklepu ve volbe
    # (napr. "--forma") mohli sami poradit nejblizsi platnou volbu, misto
    # aby argparse jen napsal genericke "unrecognized arguments" a skoncil.
    args, neznamo = ap.parse_known_args()
    if neznamo:
        known = sorted({o for a in ap._actions for o in a.option_strings})
        print("Neznama volba:")
        for u in neznamo:
            flag = u.split("=", 1)[0]
            navrh = difflib.get_close_matches(flag, known, n=1, cutoff=0.4)
            if navrh:
                print(f"  {flag!r} - mozna jsi mysel {navrh[0]!r}?")
            else:
                print(f"  {flag!r} - neznama volba")
        print("\nNapoveda: python protc_otisk.py snapshot --help")
        sys.exit(2)

    # normalizace --types: "all" = beze zmeny, jinak mnozina velkych typu (PDF, DWG, ...)
    # porovnava se proti sloupci typ_souboru, ktery uz je velkymi pismeny (viz scrape_product)
    wanted_types = None
    if args.types.strip().lower() != "all":
        wanted_types = {t.strip().upper() for t in args.types.split(",") if t.strip()}
        if not wanted_types:
            wanted_types = None

    def normalize(s):
        """Male pismena + bez diakritiky, aby '--category vzduch' naslo i
        'VZDUCH' i 'Vzduch/Voda' bez ohledu na to, jak presne to na webu je."""
        s = unicodedata.normalize("NFKD", s or "")
        s = "".join(c for c in s if not unicodedata.combining(c))
        return s.lower()

    def parse_substrings(value):
        """'a, B ,Č' -> ['a', 'b', 'c'] (normalizovano, bez prazdnych)."""
        if not value:
            return None
        parts = [normalize(p.strip()) for p in value.split(",") if p.strip()]
        return parts or None

    def matches_any(text, substrings):
        t = normalize(text)
        return any(s in t for s in substrings)

    wanted_category = parse_substrings(args.category)
    wanted_brand = parse_substrings(args.brand)
    wanted_product_type = parse_substrings(args.product_type)
    wanted_section = parse_substrings(args.section)

    run_id = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    logger, log_path = setup_logging(run_id)

    if not args.no_session and not SESSION_FILE.exists():
        logger.error("Chybi session.json - nejdriv spust:  python login.py")
        logger.error("(nebo pouzij --no-session pro otisk jen verejne casti)")
        sys.exit(1)

    # zakladni nazev bez pripony - vychozi tvar odpovida logum:
    # snapshot_RRRRMMDD_HHMMSS (rok, datum, cas, vc. sekund), ve slozce exports/
    if args.out:
        out_base = Path(args.out)
        if out_base.suffix.lower() in (".xlsx", ".csv"):
            out_base = out_base.with_suffix("")
    else:
        EXPORT_DIR.mkdir(exist_ok=True)
        out_base = EXPORT_DIR / f"snapshot_{run_id}"

    logger.info(f"Log tohoto behu: {log_path}")
    beh_t0 = time.perf_counter()

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=not args.headed)
            ctx_kwargs = {"locale": "cs-CZ"}
            if not args.no_session:
                ctx_kwargs["storage_state"] = str(SESSION_FILE)
            ctx = browser.new_context(**ctx_kwargs)
            page = ctx.new_page()

            if not args.no_session:
                logger.info("Overuji session...")
                check = check_session(page)
                if not check["ok"]:
                    logger.error(
                        f"Session vypada mrtva (stazitelnych odkazu: {check['downloads']}, "
                        f"zamcenych: {check['locked']}) - nejdriv spust:  python login.py"
                    )
                    sys.exit(1)
                logger.info(f"  session OK (stazitelnych: {check['downloads']}, zamcenych: {check['locked']})\n")

            logger.info("Ctu rozcestnik dokumentace...")
            products = scrape_index(page)

            if wanted_category is not None:
                products = [p for p in products if matches_any(p["kategorie"], wanted_category)]
            if wanted_brand is not None:
                products = [p for p in products if matches_any(p["znacka"], wanted_brand)]
            if wanted_product_type is not None:
                products = [p for p in products if matches_any(p["typ_produktu"], wanted_product_type)]

            if args.limit:
                products = products[: args.limit]
            logger.info(f"  nalezeno {len(products)} produktu / podstranek\n")

            MAX_POKUSU = 3
            records = []
            chyby = 0
            scrape_t0 = time.perf_counter()
            bar = ProgressBar()
            for i, prod in enumerate(products, start=1):
                label = prod["produkt_url"].rsplit("/", 1)[-1]
                t0 = time.perf_counter()
                recs = None
                for pokus in range(1, MAX_POKUSU + 1):
                    try:
                        recs = scrape_product(page, prod, include_nejvyhledavanejsi=args.nejvyhledavanejsi)
                        break
                    except Exception:
                        if pokus < MAX_POKUSU:
                            bar.interrupt()
                            logger.warning(
                                f"[{i:>3}/{len(products)}] {label}  -> pokus {pokus}/{MAX_POKUSU} "
                                f"selhal, zkousim znovu"
                            )
                            page.wait_for_timeout(1500)
                        else:
                            chyby += 1
                            trvani = time.perf_counter() - t0
                            bar.interrupt()
                            # exception() zaloguje i cely traceback - do log souboru,
                            # aby slo pozdeji dohledat presne v cem byl problem
                            logger.exception(
                                f"[{i:>3}/{len(products)}] {label}  -> CHYBA pri scrapovani "
                                f"po {MAX_POKUSU} pokusech  ({trvani:.1f}s)"
                            )

                if recs is not None:
                    records.extend(recs)
                    trvani = time.perf_counter() - t0
                    prumer = (time.perf_counter() - scrape_t0) / i
                    zbyva = len(products) - i
                    eta = prumer * zbyva
                    # detailni radek jde jen do log souboru (viz _ConsoleFilter) -
                    # na konzoli misto nej bezi progress bar nize
                    logger.info(
                        f"[{i:>3}/{len(products)}] {label}  -> {len(recs)} dokumentu  "
                        f"({trvani:.1f}s, ETA ~{eta:.0f}s, zbyva {zbyva})",
                        extra={"file_only": True},
                    )
                bar.update(i, len(products), extra=f"ETA ~{eta:.0f}s" if recs is not None else "")

            bar.finish()
            browser.close()
    except Exception:
        logger.exception("Beh spadl s neocekavanou chybou")
        logger.error(f"Podrobnosti (vcetne tracebacku) jsou v logu: {log_path}")
        sys.exit(1)

    if wanted_types is not None:
        pred_filtr = len(records)
        records = [r for r in records if r["typ_souboru"].upper() in wanted_types]
        logger.info(f"\nFiltr --types {args.types}: {pred_filtr} -> {len(records)} radku")

    if args.no_dwg:
        pred_filtr = len(records)
        records = [r for r in records if r["typ_souboru"].upper() != "DWG"]
        logger.info(f"Filtr --no-dwg: {pred_filtr} -> {len(records)} radku")

    if wanted_section is not None:
        pred_filtr = len(records)
        records = [r for r in records if matches_any(r["sekce"], wanted_section)]
        logger.info(f"Filtr --section {args.section}: {pred_filtr} -> {len(records)} radku")

    out_paths = []
    if args.dry_run:
        logger.info("\n--dry-run: vystup se nezapisuje")
    else:
        if args.format in ("xlsx", "both"):
            p = out_base.with_suffix(".xlsx")
            write_xlsx(records, p)
            out_paths.append(p)
        if args.format in ("csv", "both"):
            p = out_base.with_suffix(".csv")
            write_csv(records, p)
            out_paths.append(p)

    celkem_trvani = time.perf_counter() - beh_t0
    locked = sum(1 for r in records if r["dostupnost"] != "ke stazeni")
    logger.info("\nHotovo:")
    for p in out_paths:
        logger.info(f"  {p}")
    logger.info(f"  radku: {len(records)}   zamcenych: {locked}   chyb pri scrapovani: {chyby}")
    logger.info(f"  celkovy cas behu: {celkem_trvani:.1f}s")
    if locked > len(records) * 0.3:
        logger.warning("  !! Hodne zamcenych polozek - mozna vyprsela session, spust login.py znovu.")
    if chyby:
        logger.warning(f"  !! {chyby} produktu skoncilo chybou - podrobnosti (traceback) v logu: {log_path}")


if __name__ == "__main__":
    main()
